import sqlite3
import os
import openpyxl

DB_FILE = 'database.db'

# Real 2024-25 Allocation Data from PDF OCR
ms74_alt_2024_25 = {
    'HQ 55 Arty Bde': [619.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ 21 Inf Bde': [301.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ 88 Inf Bde': [762.0, 0.0, 0.0, 0.0, 0.0, 100.0, 0.0, 0.0, 0.0, 0.0, 100.0, 0.0],
    'HQ 105 Inf Bde': [619.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '9 Bengal Lancers': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 150.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '3 Engr Bn': [185.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 200.0],
    '2 Sig Bn': [249.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '5 BIR (Div Sp Bn)': [106.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '31 ST Bn': [131.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '41 Fd Amb': [93.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '71 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '505 DOC': [161.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '117 Fd Wksp Coy EME': [109.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '119 Fd Wksp Coy EME': [141.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '145 Fd Wksp Coy EME': [80.0, 0.0, 0.0, 0.0, 0.0, 0.0, 95.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 MP Unit': [198.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0],
    '55 FIU': [107.0, 400.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0, 50.0, 150.0],
    'HQ Coy 55 Inf Div': [59.0, 200.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 50.0]
}

octane100_alt_2024_25 = {
    'HQ 55 Arty Bde': [500.0, 500.0, 0.0, 300.0, 500.0, 0.0, 0.0, 0.0, 0.0, 0.0, 666.0, 200.0],
    'HQ 21 Inf Bde': [500.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 150.0, 0.0, 50.0, 195.0, 1800.0],
    'HQ 88 Inf Bde': [1000.0, 0.0, 0.0, 700.0, 700.0, 0.0, 0.0, 0.0, 100.0, 0.0, 691.0, 1700.0],
    'HQ 105 Inf Bde': [500.0, 0.0, 0.0, 500.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 501.0, 1700.0],
    '9 Bengal Lancers': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '3 Engr Bn': [0.0, 600.0, 800.0, 750.0, 3000.0, 0.0, 0.0, 0.0, 300.0, 0.0, 1000.0, 5500.0],
    '2 Sig Bn': [200.0, 400.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0],
    '5 BIR (Div Sp Bn)': [300.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '31 ST Bn': [300.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 551.0, 0.0],
    '41 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '71 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '505 DOC': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '117 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '119 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '145 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 MP Unit': [1404.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 FIU': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ Coy 55 Inf Div': [1500.0, 1917.0, 0.0, 0.0, 400.0, 0.0, 0.0, 400.0, 0.0, 0.0, 773.0, 650.0]
}

diesel_alt_2024_25 = {
    'HQ 55 Arty Bde': [16029.0, 13500.0, 23000.0, 11500.0, 13000.0, 17500.0, 17000.0, 15000.0, 11000.0, 27000.0, 17500.0, 27500.0],
    'HQ 21 Inf Bde': [13878.0, 8500.0, 11000.0, 7500.0, 11500.0, 12000.0, 7500.0, 8750.0, 14000.0, 11500.0, 6500.0, 15000.0],
    'HQ 88 Inf Bde': [14471.0, 9500.0, 7000.0, 10700.0, 9500.0, 13000.0, 12000.0, 9750.0, 10000.0, 6000.0, 10000.0, 11500.0],
    'HQ 105 Inf Bde': [25798.0, 16500.0, 7500.0, 10000.0, 9000.0, 50500.0, 0.0, 5000.0, 12000.0, 18000.0, 7000.0, 15000.0],
    '9 Bengal Lancers': [4500.0, 4000.0, 4000.0, 5500.0, 2000.0, 20000.0, 5000.0, 0.0, 3500.0, 2500.0, 4000.0, 7000.0],
    '3 Engr Bn': [11000.0, 2000.0, 600.0, 1500.0, 35000.0, 6000.0, 8500.0, 11000.0, 8500.0, 5000.0, 14000.0, 11500.0],
    '2 Sig Bn': [5966.0, 8000.0, 6500.0, 7000.0, 7000.0, 12000.0, 8500.0, 5000.0, 7500.0, 6000.0, 6000.0, 7000.0],
    '5 BIR (Div Sp Bn)': [7914.0, 8000.0, 4500.0, 12500.0, 9000.0, 23000.0, 13500.0, 18000.0, 10500.0, 12500.0, 12000.0, 14000.0],
    '31 ST Bn': [8017.0, 8000.0, 5000.0, 7500.0, 3000.0, 13000.0, 12000.0, 0.0, 11000.0, 0.0, 6000.0, 12500.0],
    '41 Fd Amb': [1300.0, 2250.0, 1500.0, 1500.0, 1000.0, 1500.0, 1250.0, 1000.0, 2200.0, 1000.0, 750.0, 2500.0],
    '71 Fd Amb': [2660.0, 1000.0, 1500.0, 2000.0, 1000.0, 750.0, 1500.0, 1500.0, 1500.0, 2250.0, 1000.0, 2500.0],
    '505 DOC': [1420.0, 1000.0, 2000.0, 500.0, 1000.0, 1000.0, 1500.0, 0.0, 1200.0, 0.0, 1000.0, 2500.0],
    '117 Fd Wksp Coy EME': [1506.0, 500.0, 1500.0, 500.0, 2000.0, 750.0, 0.0, 1000.0, 1000.0, 1000.0, 700.0, 2000.0],
    '119 Fd Wksp Coy EME': [2031.0, 1500.0, 2000.0, 500.0, 1000.0, 0.0, 1500.0, 0.0, 1500.0, 1250.0, 1500.0, 1500.0],
    '145 Fd Wksp Coy EME': [2540.0, 500.0, 0.0, 500.0, 3000.0, 0.0, 1500.0, 0.0, 0.0, 0.0, 1000.0, 1500.0],
    '55 MP Unit': [4141.0, 2500.0, 2500.0, 3000.0, 2000.0, 2000.0, 2000.0, 1500.0, 3250.0, 2000.0, 2000.0, 2500.0],
    '55 FIU': [3432.0, 1000.0, 1500.0, 500.0, 1000.0, 0.0, 1000.0, 1000.0, 0.0, 1000.0, 0.0, 2000.0],
    'HQ Coy 55 Inf Div': [5760.0, 3000.0, 2500.0, 2000.0, 2000.0, 1500.0, 2000.0, 0.0, 0.0, 3000.0, 0.0, 1873.0]
}

def seed_pol_data(conn):
    cursor = conn.cursor()
    
    # Create table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pol_monthly_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_name TEXT NOT NULL,
            fiscal_year TEXT NOT NULL,
            month TEXT NOT NULL,
            pol_grade TEXT NOT NULL,
            allocation REAL,
            expenditure REAL
        )
    ''')
    
    # Clear the table first
    cursor.execute("DELETE FROM pol_monthly_records")
    
    months = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    
    # Map for unmatched unit name
    unit_map = {
        '19 E Bengal (Div Sp Bn)': '5 BIR (Div Sp Bn)'
    }
    
    # File 1: Exp data for FY-2024-25
    file_1_path = 'scratch/file_1.xlsx'
    if os.path.exists(file_1_path):
        wb1 = openpyxl.load_workbook(file_1_path, data_only=True)
        for grade in ['Diesel', 'MS-74', '100 Octane']:
            if grade in wb1.sheetnames:
                sheet = wb1[grade]
                rows = list(sheet.iter_rows(values_only=True))
                for r in rows[4:]:
                    if not r or r[0] is None:
                        continue
                    unit = str(r[0]).strip()
                    if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                        continue
                    unit = unit_map.get(unit, unit)
                    
                    # Select the correct allocation source mapping dict
                    alt_source = None
                    if grade == 'Diesel':
                        alt_source = diesel_alt_2024_25
                    elif grade == 'MS-74':
                        alt_source = ms74_alt_2024_25
                    elif grade == '100 Octane':
                        alt_source = octane100_alt_2024_25
                    
                    for m_idx, month in enumerate(months):
                        exp_val = r[m_idx + 1]
                        if exp_val is None:
                            exp_val = 0.0
                        else:
                            try:
                                exp_val = float(exp_val)
                            except:
                                exp_val = 0.0
                        
                        # Get real allocation
                        alt_val = 0.0
                        if alt_source and unit in alt_source:
                            alt_val = alt_source[unit][m_idx]
                        
                        cursor.execute('''
                            INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (unit, '2024-25', month, grade, alt_val, exp_val))
        print("Successfully seeded FY-2024-25 records with real allocation data.")
    else:
        print("file_1.xlsx not found!")
    
    # File 2: Exp data for FY-2025-26
    # File 3: Alt data for FY-2025-26
    file_2_path = 'scratch/file_2.xlsx'
    file_3_path = 'scratch/file_3.xlsx'
    
    data_25_26 = {}
    
    if os.path.exists(file_3_path):
        wb3 = openpyxl.load_workbook(file_3_path, data_only=True)
        for grade in ['Diesel', 'MS-74', '100 Octane']:
            if grade in wb3.sheetnames:
                sheet = wb3[grade]
                rows = list(sheet.iter_rows(values_only=True))
                for r in rows[4:]:
                    if not r or r[0] is None:
                        continue
                    unit = str(r[0]).strip()
                    if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                        continue
                    unit = unit_map.get(unit, unit)
                    
                    for m_idx, month in enumerate(months):
                        alt_val = r[m_idx + 1]
                        if alt_val is None:
                            alt_val = None
                        else:
                            try:
                                alt_val = float(alt_val)
                            except:
                                alt_val = None
                        
                        key = (unit, grade, month)
                        data_25_26[key] = {'alt': alt_val, 'exp': None}
                        
    if os.path.exists(file_2_path):
        wb2 = openpyxl.load_workbook(file_2_path, data_only=True)
        for grade in ['Diesel', 'MS-74', '100 Octane']:
            if grade in wb2.sheetnames:
                sheet = wb2[grade]
                rows = list(sheet.iter_rows(values_only=True))
                for r in rows[4:]:
                    if not r or r[0] is None:
                        continue
                    unit = str(r[0]).strip()
                    if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                        continue
                    unit = unit_map.get(unit, unit)
                    
                    for m_idx, month in enumerate(months):
                        exp_val = r[m_idx + 1]
                        if exp_val is None:
                            exp_val = None
                        else:
                            try:
                                exp_val = float(exp_val)
                            except:
                                exp_val = None
                        
                        key = (unit, grade, month)
                        if key not in data_25_26:
                            data_25_26[key] = {'alt': None, 'exp': exp_val}
                        else:
                            data_25_26[key]['exp'] = exp_val

    for (unit, grade, month), vals in data_25_26.items():
        alt_val = vals['alt']
        exp_val = vals['exp']
        cursor.execute('''
            INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (unit, '2025-26', month, grade, alt_val, exp_val))
        
    conn.commit()
    print(f"Successfully seeded FY-2025-26 records. Total: {len(data_25_26)} entries.")

if __name__ == '__main__':
    conn = sqlite3.connect(DB_FILE)
    seed_pol_data(conn)
    
    # Query count
    cursor = conn.cursor()
    cursor.execute("SELECT count(*) FROM pol_monthly_records")
    print("Total records in DB:", cursor.fetchone()[0])
    
    # Show a few sample records
    cursor.execute("SELECT * FROM pol_monthly_records LIMIT 5")
    print("Samples:", cursor.fetchall())
    
    # Check total allocation for Diesel HQ 55 Arty Bde in 2024-25 (should match 209529.0)
    cursor.execute("SELECT SUM(allocation), SUM(expenditure) FROM pol_monthly_records WHERE fiscal_year = '2024-25' AND unit_name = 'HQ 55 Arty Bde' AND pol_grade = 'Diesel'")
    print("HQ 55 Arty Bde FY 2024-25 Diesel totals (Alt, Exp):", cursor.fetchone())
    
    conn.close()
