import sqlite3
import os
import random
import uuid
from flask import Flask, jsonify, request, send_from_directory

app = Flask(__name__)
DB_FILE = 'database.db'
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


BRIGADES = {
  "HQ 55 Arty Bde": ["Rawshan Ara Regt Arty", "8 Fd Regt Arty", "27 Fd Regt Arty", "Adhoc Med Bty"],
  "HQ 21 Inf Bde": ["10 E Bengal", "20 E Bengal"],
  "HQ 88 Inf Bde": ["2 E Bengal", "6 E Bengal"],
  "HQ 105 Inf Bde": ["14 BIR", "Adhoc 34 E Bengal", "37 BIR"],
  "HQ 55 Inf Div (Direct)": [
    "9 Bengal Lancers", "3 Engr Bn", "2 Sig Bn", "5 BIR (Sp Bn)", 
    "31 ST Bn", "41 Fd Amb", "71 Fd Amb", "505 DOC", 
    "117 Fd Wksp Coy EME", "119 Fd Wksp Coy EME", "145 Fd Wksp Coy EME", 
    "55 MP Unit", "55 FIU", "HQ Coy 55 Inf Div"
  ]
}

IMMUTABLE_USER_REGISTRY = [
  { "username": "aq_55_inf_div", "assigned": "HQ 55 Inf Div", "role": 6, "appointment": "AAQMG", "access": "Viewer" },
  { "username": "dq_55_inf_div", "assigned": "HQ 55 Inf Div", "role": 6, "appointment": "DAQMG", "access": "Viewer" },
  { "username": "hq_55_inf_div", "assigned": "HQ 55 Inf Div", "role": 5, "appointment": "Q Clk", "access": "Editor" },
  { "username": "dq_hq_55_arty_bde", "assigned": "HQ 55 Arty Bde", "role": 4, "appointment": "DAQMG", "access": "Viewer" },
  { "username": "hq_55_arty_bde", "assigned": "HQ 55 Arty Bde", "role": 3, "appointment": "Q Clk", "access": "Editor" },
  { "username": "dq_hq_21_inf_bde", "assigned": "HQ 21 Inf Bde", "role": 4, "appointment": "DAQMG", "access": "Viewer" },
  { "username": "hq_21_inf_bde", "assigned": "HQ 21 Inf Bde", "role": 3, "appointment": "Q Clk", "access": "Editor" },
  { "username": "dq_hq_88_inf_bde", "assigned": "HQ 88 Inf Bde", "role": 4, "appointment": "DAQMG", "access": "Viewer" },
  { "username": "hq_88_inf_bde", "assigned": "HQ 88 Inf Bde", "role": 3, "appointment": "Q Clk", "access": "Editor" },
  { "username": "dq_hq_105_inf_bde", "assigned": "HQ 105 Inf Bde", "role": 4, "appointment": "DAQMG", "access": "Viewer" },
  { "username": "hq_105_inf_bde", "assigned": "HQ 105 Inf Bde", "role": 3, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_rawshan_ara_regt", "assigned": "Rawshan Ara Regt Artly", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "rawshan_ara_regt", "assigned": "Rawshan Ara Regt Arty", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_8_fd_regt", "assigned": "8 Fd Regt Arty", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "8_fd_regt", "assigned": "8 Fd Regt Arty", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_27_fd_regt", "assigned": "27 Fd Regt Arty", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "27_fd_regt", "assigned": "27 Fd Regt Arty", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_adhoc_med_bty", "assigned": "Adhoc Med Bty", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "adhoc_med_bty", "assigned": "Adhoc Med Bty", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_10_eb", "assigned": "10 E Bengal", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "10_eb", "assigned": "10 E Bengal", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_20_eb", "assigned": "20 E Bengal", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "20_eb", "assigned": "20 E Bengal", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_6_eb", "assigned": "6 E Bengal", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "6_eb", "assigned": "6 E Bengal", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_2_eb", "assigned": "2 E Bengal", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "2_eb", "assigned": "2 E Bengal", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_14_bir", "assigned": "14 BIR", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "14_bir", "assigned": "14 BIR", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_34_eb", "assigned": "Adhoc 34 E Bengal", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "34_eb", "assigned": "Adhoc 34 E Bengal", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_37_bir", "assigned": "37 BIR", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "37_bir", "assigned": "37 BIR", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_9_bl", "assigned": "9 Bengal Lancers", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "9_bl", "assigned": "9 Bengal Lancers", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_3_engr_bn", "assigned": "3 Engr Bn", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "3_engr_bn", "assigned": "3 Engr Bn", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_2_sig_bn", "assigned": "2 Sig Bn", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "2_sig_bn", "assigned": "2 Sig Bn", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_5_bir", "assigned": "5 BIR (Sp Bn)", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "5_bir", "assigned": "5 BIR (Sp Bn)", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_31_st_bn", "assigned": "31 ST Bn", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "31_st_bn", "assigned": "31 ST Bn", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_41_fd_amb", "assigned": "41 Fd Amb", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "41_fd_amb", "assigned": "41 Fd Amb", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_71_fd_amb", "assigned": "71 Fd Amb", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "71_fd_amb", "assigned": "71 Fd Amb", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_505_doc", "assigned": "505 DOC", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "505_doc", "assigned": "505 DOC", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_117_fd_wksp", "assigned": "117 Fd Wksp Coy EME", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "117_fd_wksp", "assigned": "117 Fd Wksp Coy EME", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_119_fd_wksp", "assigned": "119 Fd Wksp Coy EME", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "119_fd_wksp", "assigned": "119 Fd Wksp Coy EME", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_145_fd_wksp", "assigned": "145 Fd Wksp Coy EME", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "145_fd_wksp", "assigned": "145 Fd Wksp Coy EME", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "qm_55_mp", "assigned": "55 MP Unit", "role": 2, "appointment": "QM", "access": "Viewer" },
  { "username": "55_mp", "assigned": "55 MP Unit", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_55_fiu", "assigned": "55 FIU", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "55_fiu", "assigned": "55 FIU", "role": 1, "appointment": "Q Clk", "access": "Editor" },
  { "username": "oc_hq_coy_55_div", "assigned": "HQ Coy 55 Inf Div", "role": 2, "appointment": "OC", "access": "Viewer" },
  { "username": "hq_coy_55_div", "assigned": "HQ Coy 55 Inf Div", "role": 1, "appointment": "Q Clk", "access": "Editor" }
]

INITIAL_LOGISTICS = {
  "HQ 55 Arty Bde": { "vAvail": 4, "vTotal": 6, "pol": 2000, "cook": 2, "waiter": 4, "strength": 80 },
  "HQ 21 Inf Bde": { "vAvail": 4, "vTotal": 6, "pol": 2200, "cook": 2, "waiter": 4, "strength": 85 },
  "HQ 88 Inf Bde": { "vAvail": 5, "vTotal": 7, "pol": 2100, "cook": 2, "waiter": 4, "strength": 90 },
  "HQ 105 Inf Bde": { "vAvail": 4, "vTotal": 6, "pol": 2300, "cook": 2, "waiter": 4, "strength": 88 },
  "Rawshan Ara Regt Arty": { "vAvail": 8, "vTotal": 12, "pol": 4500, "cook": 4, "waiter": 8, "strength": 245 },
  "8 Fd Regt Arty": { "vAvail": 6, "vTotal": 8, "pol": 3200, "cook": 3, "waiter": 6, "strength": 180 },
  "27 Fd Regt Arty": { "vAvail": 9, "vTotal": 10, "pol": 4100, "cook": 4, "waiter": 7, "strength": 215 },
  "Adhoc Med Bty": { "vAvail": 3, "vTotal": 4, "pol": 1500, "cook": 2, "waiter": 4, "strength": 90 },
  "10 E Bengal": { "vAvail": 14, "vTotal": 18, "pol": 6500, "cook": 5, "waiter": 10, "strength": 480 },
  "20 E Bengal": { "vAvail": 11, "vTotal": 15, "pol": 5800, "cook": 5, "waiter": 9, "strength": 460 },
  "2 E Bengal": { "vAvail": 12, "vTotal": 16, "pol": 6100, "cook": 5, "waiter": 10, "strength": 470 },
  "6 E Bengal": { "vAvail": 10, "vTotal": 14, "pol": 5200, "cook": 4, "waiter": 8, "strength": 430 },
  "14 BIR": { "vAvail": 15, "vTotal": 18, "pol": 6900, "cook": 6, "waiter": 11, "strength": 520 },
  "Adhoc 34 E Bengal": { "vAvail": 5, "vTotal": 8, "pol": 2800, "cook": 3, "waiter": 6, "strength": 150 },
  "37 BIR": { "vAvail": 11, "vTotal": 15, "pol": 5400, "cook": 5, "waiter": 9, "strength": 440 },
  "9 Bengal Lancers": { "vAvail": 16, "vTotal": 20, "pol": 8500, "cook": 6, "waiter": 12, "strength": 310 },
  "3 Engr Bn": { "vAvail": 22, "vTotal": 26, "pol": 9200, "cook": 5, "waiter": 10, "strength": 280 },
  "2 Sig Bn": { "vAvail": 18, "vTotal": 22, "pol": 7500, "cook": 4, "waiter": 8, "strength": 250 },
  "5 BIR (Sp Bn)": { "vAvail": 14, "vTotal": 16, "pol": 6200, "cook": 6, "waiter": 11, "strength": 490 },
  "31 ST Bn": { "vAvail": 45, "vTotal": 50, "pol": 24000, "cook": 8, "waiter": 15, "strength": 350 },
  "41 Fd Amb": { "vAvail": 12, "vTotal": 14, "pol": 3800, "cook": 3, "waiter": 6, "strength": 120 },
  "71 Fd Amb": { "vAvail": 10, "vTotal": 12, "pol": 3500, "cook": 3, "waiter": 6, "strength": 110 },
  "505 DOC": { "vAvail": 4, "vTotal": 6, "pol": 1800, "cook": 2, "waiter": 4, "strength": 60 },
  "117 Fd Wksp Coy EME": { "vAvail": 8, "vTotal": 10, "pol": 3900, "cook": 3, "waiter": 5, "strength": 140 },
  "119 Fd Wksp Coy EME": { "vAvail": 7, "vTotal": 9, "pol": 3600, "cook": 3, "waiter": 5, "strength": 130 },
  "145 Fd Wksp Coy EME": { "vAvail": 9, "vTotal": 11, "pol": 4200, "cook": 3, "waiter": 6, "strength": 150 },
  "55 MP Unit": { "vAvail": 8, "vTotal": 8, "pol": 2500, "cook": 2, "waiter": 4, "strength": 95 },
  "55 FIU": { "vAvail": 3, "vTotal": 4, "pol": 1200, "cook": 1, "waiter": 2, "strength": 45 },
  "HQ Coy 55 Inf Div": { "vAvail": 12, "vTotal": 15, "pol": 5500, "cook": 5, "waiter": 12, "strength": 220 }
}

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    # Create users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT,
            role INTEGER,
            roleName TEXT,
            appointment TEXT,
            assigned TEXT,
            assignedUnit TEXT,
            unitName TEXT,
            scopeUnit TEXT,
            scopeBde TEXT,
            is_first_login BOOLEAN,
            baNo TEXT,
            rank TEXT,
            fullName TEXT,
            mobile TEXT,
            avatar TEXT
        )
    ''')
    
    # Create logistics table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS logistics (
            unit_name TEXT PRIMARY KEY,
            vAvail INTEGER,
            vTotal INTEGER,
            pol INTEGER,
            cook INTEGER,
            waiter INTEGER,
            strength INTEGER
        )
    ''')
    
    # Create pol_monthly_records table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pol_monthly_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_name TEXT NOT NULL,
            fiscal_year TEXT NOT NULL,
            month TEXT NOT NULL,
            pol_grade TEXT NOT NULL,
            allocation REAL,
            expenditure REAL
        )
    ''')
    
    # Create pol_demands table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pol_demands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_name TEXT NOT NULL,
            fiscal_year TEXT NOT NULL,
            month TEXT NOT NULL,
            pol_grade TEXT NOT NULL,
            amount REAL NOT NULL,
            status TEXT DEFAULT 'Pending'
        )
    ''')
    
    try:
        cursor.execute("ALTER TABLE pol_demands ADD COLUMN ref_letter TEXT")
    except sqlite3.OperationalError:
        pass # Already exists
        
    conn.commit()
    
    # Seed users if empty
    cursor.execute("SELECT count(*) FROM users")
    if cursor.fetchone()[0] == 0:
        seed_users(conn)
        
    # Seed logistics if empty
    cursor.execute("SELECT count(*) FROM logistics")
    if cursor.fetchone()[0] == 0:
        seed_logistics(conn)
        
    # Seed pol_monthly_records if empty
    cursor.execute("SELECT count(*) FROM pol_monthly_records")
    if cursor.fetchone()[0] == 0:
        seed_pol_data(conn)
        
    # Patch database: check if 'HQ 55 Inf Div' is in logistics
    cursor.execute("SELECT count(*) FROM logistics WHERE unit_name = 'HQ 55 Inf Div'")
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO logistics (unit_name, vAvail, vTotal, pol, cook, waiter, strength)
            VALUES (?, 0, 0, ?, 0, 0, 0)
        ''', ('HQ 55 Inf Div', 209097))
        conn.commit()

    # Patch database: check if 'HQ 55 Inf Div' has records in pol_monthly_records
    cursor.execute("SELECT count(*) FROM pol_monthly_records WHERE unit_name = 'HQ 55 Inf Div'")
    if cursor.fetchone()[0] == 0:
        div_allocations = {
            'Diesel': 1292544.0,
            'MS-74': 4737.0,
            '100 Octane': 26685.0
        }
        for grade, amount in div_allocations.items():
            cursor.execute('''
                INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', ('HQ 55 Inf Div', '2025-26', 'Jul', grade, amount, 0.0))
            for month in ['Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']:
                cursor.execute('''
                    INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', ('HQ 55 Inf Div', '2025-26', month, grade, 0.0, 0.0))
        conn.commit()

    # Rename usernames for Brigade DQ users if they exist in the database
    old_to_new_users = {
        "dq_55_arty_bde": "dq_hq_55_arty_bde",
        "dq_21_inf_bde": "dq_hq_21_inf_bde",
        "dq_88_inf_bde": "dq_hq_88_inf_bde",
        "dq_105_inf_bde": "dq_hq_105_inf_bde"
    }
    for old_u, new_u in old_to_new_users.items():
        cursor.execute("SELECT count(*) FROM users WHERE username = ?", (old_u,))
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT count(*) FROM users WHERE username = ?", (new_u,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("UPDATE users SET username = ? WHERE username = ?", (new_u, old_u))
                
    # Patch database: check if brigade HQs are in logistics
    bde_logistics = {
        "HQ 55 Arty Bde": { "vAvail": 4, "vTotal": 6, "pol": 2000, "cook": 2, "waiter": 4, "strength": 80 },
        "HQ 21 Inf Bde": { "vAvail": 4, "vTotal": 6, "pol": 2200, "cook": 2, "waiter": 4, "strength": 85 },
        "HQ 88 Inf Bde": { "vAvail": 5, "vTotal": 7, "pol": 2100, "cook": 2, "waiter": 4, "strength": 90 },
        "HQ 105 Inf Bde": { "vAvail": 4, "vTotal": 6, "pol": 2300, "cook": 2, "waiter": 4, "strength": 88 }
    }
    for bde_name, data in bde_logistics.items():
        cursor.execute("SELECT count(*) FROM logistics WHERE unit_name = ?", (bde_name,))
        if cursor.fetchone()[0] == 0:
            cursor.execute('''
                INSERT INTO logistics (unit_name, vAvail, vTotal, pol, cook, waiter, strength)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (bde_name, data["vAvail"], data["vTotal"], data["pol"], data["cook"], data["waiter"], data["strength"]))
            
    conn.commit()
        
    conn.close()

def seed_users(conn):
    cursor = conn.cursor()
    for u in IMMUTABLE_USER_REGISTRY:
        scopeUnit = None
        scopeBde = None
        roleName = ""
        defaultRank = ""
        
        assignedName = u["assigned"]
        if assignedName == "Rawshan Ara Regt Artly":
            assignedName = "Rawshan Ara Regt Arty"
            
        if u["role"] == 1 or u["role"] == 2:
            scopeUnit = assignedName
            scopeBde = next((b for b in BRIGADES if assignedName in BRIGADES[b]), None)
            roleName = "Unit Clerk" if u["role"] == 1 else "Unit QM / OC"
            defaultRank = "Warrant Officer" if u["role"] == 1 else "Major"
        elif u["role"] == 3 or u["role"] == 4:
            scopeBde = assignedName
            roleName = "Brigade Clerk" if u["role"] == 3 else "Brigade DAQMG"
            defaultRank = "Sergeant" if u["role"] == 3 else "Major"
        elif u["role"] == 5 or u["role"] == 6:
            scopeBde = "ALL"
            roleName = "Division Q Clerk" if u["role"] == 5 else ("Division AAQMG" if u["appointment"] == "AAQMG" else "Division DAQMG")
            defaultRank = "Senior Warrant Officer" if u["role"] == 5 else ("Lieutenant Colonel" if u["appointment"] == "AAQMG" else "Major")
            
        cursor.execute('''
            INSERT INTO users (username, password, role, roleName, appointment, assigned, assignedUnit, unitName, scopeUnit, scopeBde, is_first_login, baNo, rank, fullName, mobile, avatar)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            u["username"], "temp123", u["role"], roleName, u["appointment"], assignedName, assignedName, assignedName, scopeUnit, scopeBde,
            True, "", defaultRank, f"{u['appointment']} ({assignedName})", "", f"https://avatar.iran.liara.run/public/{random.randint(1, 70)}"
        ))
    conn.commit()

def seed_logistics(conn):
    cursor = conn.cursor()
    for unit, data in INITIAL_LOGISTICS.items():
        cursor.execute('''
            INSERT INTO logistics (unit_name, vAvail, vTotal, pol, cook, waiter, strength)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (unit, data["vAvail"], data["vTotal"], data["pol"], data["cook"], data["waiter"], data["strength"]))
    conn.commit()

# --- Flask Web Routes ---

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def serve_file(path):
    res = send_from_directory('.', path)
    if path.endswith('.js'):
        res.headers['Content-Type'] = 'text/javascript'
    return res

# --- API Endpoints ---

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username_or_mobile = data.get("username", "").strip().lower()
    password = data.get("password", "")
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Try username lookup first
    cursor.execute("SELECT * FROM users WHERE LOWER(username) = ?", (username_or_mobile,))
    user_row = cursor.fetchone()
    
    if not user_row:
        # If not found, look up all users and check their mobile numbers
        cursor.execute("SELECT * FROM users")
        all_users = cursor.fetchall()
        for u in all_users:
            mobile = u["mobile"] or ""
            # Strip non-digits from both mobile and input
            clean_mobile = ''.join(c for c in mobile if c.isdigit())
            clean_input = ''.join(c for c in username_or_mobile if c.isdigit())
            if clean_input and (clean_mobile == clean_input or mobile == username_or_mobile or (len(mobile) >= 4 and mobile[-4:] == username_or_mobile)):
                user_row = u
                break
                
    conn.close()
    
    if user_row:
        if user_row["password"] == password:
            user_dict = dict(user_row)
            user_dict["is_first_login"] = bool(user_dict["is_first_login"])
            return jsonify({"success": True, "user": user_dict})
        else:
            is_first_login = bool(user_row["is_first_login"])
            if password == "temp123" and not is_first_login:
                return jsonify({"success": False, "message": "You entered an old password.", "old_password": True}), 401
            return jsonify({"success": False, "message": "Incorrect password."}), 401
            
    return jsonify({"success": False, "message": "Incorrect username or phone."}), 401

@app.route('/api/impersonate', methods=['POST'])
def impersonate():
    data = request.json
    username = data.get("username", "").strip().lower()
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
    user_row = cursor.fetchone()
    conn.close()
    
    if user_row:
        user_dict = dict(user_row)
        user_dict["is_first_login"] = bool(user_dict["is_first_login"])
        return jsonify({"success": True, "user": user_dict})
    
    return jsonify({"success": False, "message": "User not found"}), 404

@app.route('/api/users', methods=['GET'])
def get_users():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users")
    rows = cursor.fetchall()
    conn.close()
    
    users_data = {}
    for row in rows:
        user_dict = dict(row)
        user_dict["is_first_login"] = bool(user_dict["is_first_login"])
        users_data[row["username"]] = user_dict
    return jsonify(users_data)

@app.route('/api/logistics', methods=['GET'])
def get_logistics():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM logistics")
    rows = cursor.fetchall()
    conn.close()
    
    logistics_data = {}
    for row in rows:
        logistics_data[row["unit_name"]] = {
            "vAvail": row["vAvail"],
            "vTotal": row["vTotal"],
            "pol": row["pol"],
            "cook": row["cook"],
            "waiter": row["waiter"],
            "strength": row["strength"]
        }
    return jsonify(logistics_data)

@app.route('/api/logistics', methods=['POST'])
def update_logistics():
    data = request.json
    unit_name = data.get("unitName")
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE logistics 
        SET vAvail = ?, vTotal = ?, pol = ?, cook = ?, waiter = ?, strength = ?
        WHERE unit_name = ?
    ''', (data["vAvail"], data["vTotal"], data["pol"], data["cook"], data["waiter"], data["strength"], unit_name))
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route('/api/profile', methods=['POST'])
def update_profile():
    data = request.json
    username = data.get("username")
    mobile = data.get("mobile")
    
    conn = get_db()
    cursor = conn.cursor()
    
    if mobile is not None:
        cursor.execute('''
            UPDATE users
            SET rank = ?, baNo = ?, fullName = ?, avatar = ?, mobile = ?
            WHERE username = ?
        ''', (data["rank"], data["baNo"], data["fullName"], data["avatar"], mobile, username))
    else:
        cursor.execute('''
            UPDATE users
            SET rank = ?, baNo = ?, fullName = ?, avatar = ?
            WHERE username = ?
        ''', (data["rank"], data["baNo"], data["fullName"], data["avatar"], username))
        
    conn.commit()
    
    cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
    user_row = cursor.fetchone()
    conn.close()
    
    user_dict = dict(user_row)
    user_dict["is_first_login"] = bool(user_dict["is_first_login"])
    return jsonify({"success": True, "user": user_dict})

@app.route('/api/change-password', methods=['POST'])
def change_password():
    data = request.json
    username = data.get("username")
    new_password = data.get("newPassword")
    is_first_login = data.get("is_first_login", False)
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE users
        SET password = ?, is_first_login = ?
        WHERE username = ?
    ''', (new_password, is_first_login, username))
    conn.commit()
    
    cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
    user_row = cursor.fetchone()
    conn.close()
    
    user_dict = dict(user_row)
    user_dict["is_first_login"] = bool(user_dict["is_first_login"])
    return jsonify({"success": True, "user": user_dict})

@app.route('/api/reset', methods=['POST'])
def reset_database():
    if os.path.exists(DB_FILE):
        os.remove(DB_FILE)
    init_db()
    return jsonify({"success": True})

# --- Real POL Data Ingestion and Seeding ---
LINE_MOCK_DATA = {
    "2023-24": {
        "Diesel": { "total": 1029000, "last": 85000, "alt": 1060000, "data": [85000, 80000, 88000, 85000, 82000, 85000, 90000, 94000, 88000, 86000, 84000, 92000], "altData": [90000, 82000, 84000, 89000, 85000, 82000, 93000, 90000, 91000, 84000, 88000, 102000] },
        "MS-74": { "total": 399000, "last": 35000, "alt": 413000, "data": [32000, 30000, 34000, 33000, 31000, 32000, 35000, 37000, 34000, 33000, 32000, 36000], "altData": [34000, 30000, 32000, 35000, 33000, 31000, 38000, 35000, 36000, 31000, 34000, 44000] },
        "100 Octane": { "total": 174000, "last": 15000, "alt": 180000, "data": [14000, 13000, 15000, 14000, 13000, 14000, 16000, 17000, 15000, 14000, 13000, 16000], "altData": [15000, 12000, 14000, 16000, 15000, 13000, 17000, 15000, 16000, 13000, 15000, 19000] }
    },
    "2026-27": {
        "Diesel": { "total": 1200000, "last": 95000, "alt": 1220000, "data": [85000, 83000, 94000, 92000, 90000, 93000, 98000, 102000, 96000, 94000, 92000, 99000], "altData": [90000, 80000, 90000, 95000, 92000, 89000, 100000, 98000, 99000, 90000, 95000, 112000] },
        "MS-74": { "total": 500000, "last": 42000, "alt": 510000, "data": [36000, 35000, 40000, 39000, 37000, 38000, 42000, 44000, 40000, 39000, 38000, 43000], "altData": [40000, 33000, 38000, 41000, 36000, 40000, 44000, 42000, 43000, 38000, 40000, 75000] },
        "100 Octane": { "total": 220000, "last": 19000, "alt": 230000, "data": [17000, 16000, 18000, 17000, 16000, 17000, 19000, 20000, 18000, 17000, 16000, 19000], "altData": [19000, 15000, 17000, 19000, 15000, 18000, 20000, 18000, 19000, 16000, 17000, 27000] }
    }
}

# Real 2024-25 Allocation Data from PDF OCR
ms74_alt_2024_25 = {
    'HQ 55 Arty Bde': [619.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ 21 Inf Bde': [301.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ 88 Inf Bde': [762.0, 0.0, 0.0, 0.0, 0.0, 100.0, 0.0, 0.0, 0.0, 0.0, 100.0, 0.0],
    'HQ 105 Inf Bde': [619.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '9 Bengal Lancers': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 150.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '3 Engr Bn': [185.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 200.0],
    '2 Sig Bn': [249.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '19 E Bengal (Sp Bn)': [106.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '31 ST Bn': [131.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '41 Fd Amb': [93.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '71 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '505 DOC': [161.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '117 Fd Wksp Coy EME': [109.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '119 Fd Wksp Coy EME': [141.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '145 Fd Wksp Coy EME': [80.0, 0.0, 0.0, 0.0, 0.0, 0.0, 95.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 MP Unit': [198.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0],
    '55 FIU': [107.0, 400.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0, 50.0, 150.0],
    'HQ Coy 55 Inf Div': [59.0, 200.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 50.0]
}

octane100_alt_2024_25 = {
    'HQ 55 Arty Bde': [500.0, 500.0, 0.0, 300.0, 500.0, 0.0, 0.0, 0.0, 0.0, 0.0, 666.0, 200.0],
    'HQ 21 Inf Bde': [500.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 150.0, 0.0, 50.0, 195.0, 1800.0],
    'HQ 88 Inf Bde': [1000.0, 0.0, 0.0, 700.0, 700.0, 0.0, 0.0, 0.0, 100.0, 0.0, 691.0, 1700.0],
    'HQ 105 Inf Bde': [500.0, 0.0, 0.0, 500.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 501.0, 1700.0],
    '9 Bengal Lancers': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '3 Engr Bn': [0.0, 600.0, 800.0, 750.0, 3000.0, 0.0, 0.0, 0.0, 300.0, 0.0, 1000.0, 5500.0],
    '2 Sig Bn': [200.0, 400.0, 0.0, 200.0, 0.0, 0.0, 0.0, 200.0, 0.0, 0.0, 0.0, 0.0],
    '19 E Bengal (Sp Bn)': [300.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '31 ST Bn': [300.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 551.0, 0.0],
    '41 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '71 Fd Amb': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '505 DOC': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '117 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '119 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '145 Fd Wksp Coy EME': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 MP Unit': [1404.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    '55 FIU': [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
    'HQ Coy 55 Inf Div': [1500.0, 1917.0, 0.0, 0.0, 400.0, 0.0, 0.0, 400.0, 0.0, 0.0, 773.0, 650.0]
}

diesel_alt_2024_25 = {
    'HQ 55 Arty Bde': [16029.0, 13500.0, 23000.0, 11500.0, 13000.0, 17500.0, 17000.0, 15000.0, 11000.0, 27000.0, 17500.0, 27500.0],
    'HQ 21 Inf Bde': [13878.0, 8500.0, 11000.0, 7500.0, 11500.0, 12000.0, 7500.0, 8750.0, 14000.0, 11500.0, 6500.0, 1500.0],
    'HQ 88 Inf Bde': [14471.0, 9500.0, 7000.0, 10700.0, 9500.0, 13000.0, 12000.0, 9750.0, 10000.0, 6000.0, 10000.0, 11500.0],
    'HQ 105 Inf Bde': [25798.0, 16500.0, 7500.0, 10000.0, 9000.0, 50500.0, 0.0, 5000.0, 12000.0, 18000.0, 7000.0, 15000.0],
    '9 Bengal Lancers': [4500.0, 4000.0, 4000.0, 5500.0, 2000.0, 20000.0, 5000.0, 0.0, 3500.0, 2500.0, 4000.0, 7000.0],
    '3 Engr Bn': [11000.0, 2000.0, 600.0, 1500.0, 35000.0, 6000.0, 8500.0, 11000.0, 8500.0, 5000.0, 14000.0, 11500.0],
    '2 Sig Bn': [5966.0, 8000.0, 6500.0, 7000.0, 7000.0, 12000.0, 8500.0, 5000.0, 7500.0, 6000.0, 6000.0, 7000.0],
    '19 E Bengal (Sp Bn)': [7914.0, 8000.0, 4500.0, 12500.0, 9000.0, 23000.0, 13500.0, 18000.0, 10500.0, 12500.0, 12000.0, 14000.0],
    '31 ST Bn': [8017.0, 8000.0, 5000.0, 7500.0, 3000.0, 13000.0, 12000.0, 0.0, 11000.0, 0.0, 6000.0, 12500.0],
    '41 Fd Amb': [1300.0, 2250.0, 1500.0, 1500.0, 1000.0, 1500.0, 1250.0, 1000.0, 2200.0, 1000.0, 750.0, 2500.0],
    '71 Fd Amb': [2660.0, 1000.0, 1500.0, 2000.0, 1000.0, 750.0, 1500.0, 1500.0, 1500.0, 2250.0, 1000.0, 2500.0],
    '505 DOC': [1420.0, 1000.0, 2000.0, 500.0, 1000.0, 1000.0, 1500.0, 0.0, 1200.0, 0.0, 1000.0, 2500.0],
    '117 Fd Wksp Coy EME': [1506.0, 500.0, 1500.0, 500.0, 2000.0, 750.0, 0.0, 1000.0, 1000.0, 1000.0, 700.0, 2000.0],
    '119 Fd Wksp Coy EME': [2031.0, 1500.0, 2000.0, 500.0, 1000.0, 0.0, 1500.0, 0.0, 1500.0, 1250.0, 1500.0, 1500.0],
    '145 Fd Wksp Coy EME': [2540.0, 500.0, 0.0, 500.0, 3000.0, 0.0, 1500.0, 0.0, 0.0, 0.0, 1000.0, 1500.0],
    '55 MP Unit': [4141.0, 2500.0, 2500.0, 3000.0, 2000.0, 2000.0, 2000.0, 1500.0, 3250.0, 2000.0, 2000.0, 2500.0],
    '55 FIU': [3432.0, 1000.0, 1500.0, 500.0, 1000.0, 0.0, 1000.0, 1000.0, 0.0, 1000.0, 0.0, 2000.0],
    'HQ Coy 55 Inf Div': [5760.0, 3000.0, 2500.0, 2000.0, 2000.0, 1500.0, 2000.0, 0.0, 0.0, 3000.0, 0.0, 1873.0]
}

def seed_pol_data(conn):
    import openpyxl
    cursor = conn.cursor()
    months = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    unit_map = {
        '19 E Bengal (Div Sp Bn)': '19 E Bengal (Sp Bn)',
        '5 BIR (Div Sp Bn)': '19 E Bengal (Sp Bn)'
    }
    
    # 1. FY 2024-25 Exp
    file_1_path = 'scratch/file_1.xlsx'
    if os.path.exists(file_1_path):
        try:
            wb1 = openpyxl.load_workbook(file_1_path, data_only=True)
            for grade in ['Diesel', 'MS-74', '100 Octane']:
                if grade in wb1.sheetnames:
                    sheet = wb1[grade]
                    rows = list(sheet.iter_rows(values_only=True))
                    for r in rows[4:]:
                        if not r or r[0] is None:
                            continue
                        unit = str(r[0]).strip()
                        if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                            continue
                        unit = unit_map.get(unit, unit)
                        
                        # Find correct allocation source
                        alt_source = None
                        if grade == 'Diesel':
                            alt_source = diesel_alt_2024_25
                        elif grade == 'MS-74':
                            alt_source = ms74_alt_2024_25
                        elif grade == '100 Octane':
                            alt_source = octane100_alt_2024_25
                        
                        for m_idx, month in enumerate(months):
                            exp_val = r[m_idx + 1]
                            if exp_val is None:
                                exp_val = 0.0
                            else:
                                try:
                                    exp_val = float(exp_val)
                                except:
                                    exp_val = 0.0
                            
                            alt_val = 0.0
                            if alt_source and unit in alt_source:
                                alt_val = alt_source[unit][m_idx]
                            
                            cursor.execute('''
                                INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (unit, '2024-25', month, grade, alt_val, exp_val))
        except Exception as e:
            print("Failed seeding FY 2024-25:", e)
            
    # 2. FY 2025-26 Exp and Alt
    file_2_path = 'scratch/file_2.xlsx'
    file_3_path = 'scratch/file_3.xlsx'
    
    data_25_26 = {}
    
    try:
        if os.path.exists(file_3_path):
            wb3 = openpyxl.load_workbook(file_3_path, data_only=True)
            for grade in ['Diesel', 'MS-74', '100 Octane']:
                if grade in wb3.sheetnames:
                    sheet = wb3[grade]
                    rows = list(sheet.iter_rows(values_only=True))
                    for r in rows[4:]:
                        if not r or r[0] is None:
                            continue
                        unit = str(r[0]).strip()
                        if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                            continue
                        unit = unit_map.get(unit, unit)
                        
                        for m_idx, month in enumerate(months):
                            alt_val = r[m_idx + 1]
                            if alt_val is None:
                                alt_val = None
                            else:
                                try:
                                    alt_val = float(alt_val)
                                except:
                                    alt_val = None
                            
                            key = (unit, grade, month)
                            data_25_26[key] = {'alt': alt_val, 'exp': None}
                            
        if os.path.exists(file_2_path):
            wb2 = openpyxl.load_workbook(file_2_path, data_only=True)
            for grade in ['Diesel', 'MS-74', '100 Octane']:
                if grade in wb2.sheetnames:
                    sheet = wb2[grade]
                    rows = list(sheet.iter_rows(values_only=True))
                    for r in rows[4:]:
                        if not r or r[0] is None:
                            continue
                        unit = str(r[0]).strip()
                        if "total" in unit.lower() or "grand" in unit.lower() or unit == "":
                            continue
                        unit = unit_map.get(unit, unit)
                        
                        for m_idx, month in enumerate(months):
                            exp_val = r[m_idx + 1]
                            if exp_val is None:
                                exp_val = None
                            else:
                                try:
                                    exp_val = float(exp_val)
                                except:
                                    exp_val = None
                            
                            key = (unit, grade, month)
                            if key not in data_25_26:
                                data_25_26[key] = {'alt': None, 'exp': exp_val}
                            else:
                                data_25_26[key]['exp'] = exp_val
                                
        for (unit, grade, month), vals in data_25_26.items():
            cursor.execute('''
                INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (unit, '2025-26', month, grade, vals['alt'], vals['exp']))
    except Exception as e:
        print("Failed seeding FY 2025-26:", e)
        
    conn.commit()

@app.route('/api/pol/summary', methods=['GET'])
def get_pol_summary():
    year = request.args.get("year", "2025-26")
    grade = request.args.get("grade", "Diesel")
    scope = request.args.get("scope", "division")
    assigned = request.args.get("assigned", "")
    
    # Return all zeros for other years (such as 2026-27) as requested
    if year not in ['2024-25', '2025-26']:
        return jsonify({
            "altData": [0.0] * 12,
            "data": [0.0] * 12,
            "alt": 0.0,
            "total": 0.0
        })
        
    conn = get_db()
    cursor = conn.cursor()
    
    months = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    
    # If the year is 2025-26, show up to February 2026 as March data is not ready yet
    if year == '2025-26':
        active_months = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb']
    else:
        active_months = months
        
    units_to_query = []
    if scope == "unit":
        units_to_query = [assigned]
    elif scope == "brigade":
        # Sum both the brigade HQ itself and its subordinate units
        units_to_query = [assigned] + BRIGADES.get(assigned, [])
        
    if year in ['2024-25', '2025-26']:
        units_to_query = [('19 E Bengal (Sp Bn)' if u == '5 BIR (Sp Bn)' else u) for u in units_to_query]
        
    alt_data = []
    exp_data = []
    
    for month in active_months:
        if scope == "division":
            cursor.execute('''
                SELECT SUM(allocation), SUM(expenditure) 
                FROM pol_monthly_records 
                WHERE fiscal_year = ? AND month = ? AND pol_grade = ? AND unit_name != 'HQ 55 Inf Div'
            ''', (year, month, grade))
        else:
            placeholders = ','.join('?' for _ in units_to_query)
            cursor.execute(f'''
                SELECT SUM(allocation), SUM(expenditure) 
                FROM pol_monthly_records 
                WHERE fiscal_year = ? AND month = ? AND pol_grade = ? AND unit_name IN ({placeholders})
            ''', (year, month, grade, *units_to_query))
            
        row = cursor.fetchone()
        alt_val = row[0] if row and row[0] is not None else None
        exp_val = row[1] if row and row[1] is not None else None
        
        alt_data.append(alt_val)
        exp_data.append(exp_val)
        
    total_alt = sum(v for v in alt_data if v is not None)
    total_exp = sum(v for v in exp_data if v is not None)
    
    conn.close()
    
    return jsonify({
        "altData": alt_data,
        "data": exp_data,
        "alt": total_alt,
        "total": total_exp
    })

@app.route('/api/pol/state', methods=['GET'])
def get_pol_state():
    role = int(request.args.get("role", 6))
    assigned = request.args.get("assigned", "")
    bde = request.args.get("bde", "")
    
    conn = get_db()
    cursor = conn.cursor()
    
    grades = ['Diesel', 'MS-74', '100 Octane']
    result = []
    
    # Division users
    if role in [5, 6]:
        for grade in grades:
            # col1: Alt from Area = sum of allocations to 'HQ 55 Inf Div'
            cursor.execute('''
                SELECT SUM(allocation) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name = 'HQ 55 Inf Div'
            ''', (grade,))
            col1 = cursor.fetchone()[0] or 0.0
            
            # col2: Alt to Bde/Unit = sum of allocations to other units
            cursor.execute('''
                SELECT SUM(allocation) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name != 'HQ 55 Inf Div'
            ''', (grade,))
            col2 = cursor.fetchone()[0] or 0.0
            
            result.append({
                "grade": grade,
                "col1": col1,
                "col2": col2,
                "bal": col1 - col2
            })
            
    # Brigade users
    elif role in [3, 4]:
        target_bde = assigned or bde
        units_to_query = [target_bde] + BRIGADES.get(target_bde, [])
        placeholders = ','.join('?' for _ in units_to_query)
        
        for grade in grades:
            # col1: Total Alt
            cursor.execute(f'''
                SELECT SUM(allocation) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name IN ({placeholders})
            ''', (grade, *units_to_query))
            col1 = cursor.fetchone()[0] or 0.0
            
            # col2: Total Exp
            cursor.execute(f'''
                SELECT SUM(expenditure) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name IN ({placeholders})
            ''', (grade, *units_to_query))
            col2 = cursor.fetchone()[0] or 0.0
            
            result.append({
                "grade": grade,
                "col1": col1,
                "col2": col2,
                "bal": col1 - col2
            })
            
    # Unit users (role 1, 2)
    else:
        for grade in grades:
            # col1: Total Alt
            cursor.execute('''
                SELECT SUM(allocation) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name = ?
            ''', (grade, assigned))
            col1 = cursor.fetchone()[0] or 0.0
            
            # col2: Total Exp
            cursor.execute('''
                SELECT SUM(expenditure) FROM pol_monthly_records 
                WHERE fiscal_year = '2025-26' AND pol_grade = ? AND unit_name = ?
            ''', (grade, assigned))
            col2 = cursor.fetchone()[0] or 0.0
            
            result.append({
                "grade": grade,
                "col1": col1,
                "col2": col2,
                "bal": col1 - col2
            })
            
    conn.close()
    return jsonify(result)

@app.route('/uploads/<path:filename>')
def serve_uploads(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/api/pol/upload_ref', methods=['POST'])
def upload_ref_letter():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No selected file"}), 400
    if file:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext in ['.pdf', '.png', '.jpg', '.jpeg']:
            filename = f"{uuid.uuid4()}{ext}"
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            return jsonify({"success": True, "filename": filename})
        else:
            return jsonify({"success": False, "error": "Invalid file type. Only PDF, PNG, JPEG allowed."}), 400

@app.route('/api/pol/unit_metrics', methods=['GET'])
def get_pol_unit_metrics():
    unit_name = request.args.get("unitName")
    pol_grade = request.args.get("polGrade")
    fiscal_year = request.args.get("fiscalYear", "2025-26")
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Calculate monthly average and balance
    cursor.execute('''
        SELECT SUM(allocation), SUM(expenditure), COUNT(month) 
        FROM pol_monthly_records 
        WHERE unit_name = ? AND pol_grade = ? AND fiscal_year = ?
    ''', (unit_name, pol_grade, fiscal_year))
    stats_row = cursor.fetchone()
    
    total_alloc = stats_row[0] or 0.0
    total_exp = stats_row[1] or 0.0
    months_count = stats_row[2] or 12
    if months_count == 0:
        months_count = 12
        
    avg_exp = total_exp / months_count
    bal = total_alloc - total_exp
    
    if avg_exp > 0:
        days_last = int(round((bal * 30.0) / avg_exp))
    else:
        days_last = 0
        
    conn.close()
    return jsonify({
        "avgExp": avg_exp,
        "bal": bal,
        "daysLast": days_last
    })

@app.route('/api/pol/demand', methods=['POST'])
def add_pol_demand():
    data = request.json
    unit_name = data.get("unitName")
    month = data.get("month")
    fiscal_year = data.get("fiscalYear", "2025-26")
    pol_grade = data.get("polGrade")
    amount = float(data.get("amount", 0))
    ref_letter = data.get("refLetter", "")
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id FROM pol_demands 
        WHERE unit_name = ? AND fiscal_year = ? AND month = ? AND pol_grade = ?
    ''', (unit_name, fiscal_year, month, pol_grade))
    row = cursor.fetchone()
    
    if row:
        cursor.execute('''
            UPDATE pol_demands SET amount = ?, ref_letter = ?, status = 'Pending'
            WHERE id = ?
        ''', (amount, ref_letter, row[0]))
    else:
        cursor.execute('''
            INSERT INTO pol_demands (unit_name, fiscal_year, month, pol_grade, amount, ref_letter, status)
            VALUES (?, ?, ?, ?, ?, ?, 'Pending')
        ''', (unit_name, fiscal_year, month, pol_grade, amount, ref_letter))
        
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/pol/demands', methods=['GET'])
def get_pol_demands():
    role = int(request.args.get("role", 6))
    assigned = request.args.get("assigned", "")
    
    conn = get_db()
    cursor = conn.cursor()
    
    if role in [1, 2]:
        cursor.execute('''
            SELECT * FROM pol_demands 
            WHERE unit_name = ? ORDER BY id DESC
        ''', (assigned,))
    elif role in [3, 4]:
        sub_units = [assigned] + BRIGADES.get(assigned, [])
        placeholders = ','.join('?' for _ in sub_units)
        cursor.execute(f'''
            SELECT * FROM pol_demands 
            WHERE unit_name IN ({placeholders}) ORDER BY id DESC
        ''', (*sub_units,))
    else:
        cursor.execute('SELECT * FROM pol_demands ORDER BY id DESC')
        
    rows = cursor.fetchall()
    
    demands = []
    for r in rows:
        unit_name = r["unit_name"]
        pol_grade = r["pol_grade"]
        fiscal_year = r["fiscal_year"]
        
        # Calculate monthly average and balance for target unit
        cursor.execute('''
            SELECT SUM(allocation), SUM(expenditure), COUNT(month) 
            FROM pol_monthly_records 
            WHERE unit_name = ? AND pol_grade = ? AND fiscal_year = ?
        ''', (unit_name, pol_grade, fiscal_year))
        stats_row = cursor.fetchone()
        
        total_alloc = stats_row[0] or 0.0
        total_exp = stats_row[1] or 0.0
        months_count = stats_row[2] or 12
        if months_count == 0:
            months_count = 12
            
        avg_exp = total_exp / months_count
        bal = total_alloc - total_exp
        
        if avg_exp > 0:
            days_last = int(round((bal * 30.0) / avg_exp))
        else:
            days_last = 0
            
        # Calculate AQ Branch (Division HQ) balance
        cursor.execute('''
            SELECT SUM(allocation) - SUM(expenditure) 
            FROM pol_monthly_records 
            WHERE unit_name = 'HQ 55 Inf Div' AND pol_grade = ? AND fiscal_year = ?
        ''', (pol_grade, fiscal_year))
        aq_row = cursor.fetchone()
        aq_bal = aq_row[0] if aq_row and aq_row[0] is not None else 0.0
        
        demands.append({
            "id": r["id"],
            "unitName": r["unit_name"],
            "fiscalYear": r["fiscal_year"],
            "month": r["month"],
            "polGrade": r["pol_grade"],
            "amount": r["amount"],
            "status": r["status"],
            "refLetter": r["ref_letter"] or "",
            "avgExp": avg_exp,
            "bal": bal,
            "daysLast": days_last,
            "aqBal": aq_bal
        })
        
    conn.close()
    return jsonify(demands)

@app.route('/api/pol/allocate_demand', methods=['POST'])
def allocate_pol_demand():
    data = request.json
    demand_id = int(data.get("demandId"))
    amount = float(data.get("amount", 0))
    from_entity = data.get("fromEntity")
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Fetch demand
    cursor.execute("SELECT * FROM pol_demands WHERE id = ?", (demand_id,))
    demand = cursor.fetchone()
    if not demand:
        conn.close()
        return jsonify({"success": False, "error": "Demand not found"}), 404
        
    to_entity = demand["unit_name"]
    month = demand["month"]
    fiscal_year = demand["fiscal_year"]
    pol_grade = demand["pol_grade"]
    
    # 1. Update pol_monthly_records for target unit (to_entity)
    cursor.execute('''
        SELECT id, allocation FROM pol_monthly_records 
        WHERE unit_name = ? AND fiscal_year = ? AND month = ? AND pol_grade = ?
    ''', (to_entity, fiscal_year, month, pol_grade))
    row = cursor.fetchone()
    if row:
        current_alloc = row[1] or 0.0
        cursor.execute('''
            UPDATE pol_monthly_records SET allocation = ? 
            WHERE id = ?
        ''', (current_alloc + amount, row[0]))
    else:
        cursor.execute('''
            INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
            VALUES (?, ?, ?, ?, ?, 0.0)
        ''', (to_entity, fiscal_year, month, pol_grade, amount))
        
    # 2. Update logistics.pol for target unit
    cursor.execute("SELECT pol FROM logistics WHERE unit_name = ?", (to_entity,))
    target_row = cursor.fetchone()
    if target_row:
        cursor.execute('''
            UPDATE logistics SET pol = ? WHERE unit_name = ?
        ''', (target_row[0] + int(amount), to_entity))
    else:
        cursor.execute('''
            INSERT INTO logistics (unit_name, vAvail, vTotal, pol, cook, waiter, strength)
            VALUES (?, 0, 0, ?, 0, 0, 0)
        ''', (to_entity, int(amount)))
        
    # 3. Deduct from source unit (from_entity) if it's not 'Area HQ'
    if from_entity != 'Area HQ':
        cursor.execute("SELECT pol FROM logistics WHERE unit_name = ?", (from_entity,))
        source_row = cursor.fetchone()
        if source_row:
            cursor.execute('''
                UPDATE logistics SET pol = ? WHERE unit_name = ?
            ''', (max(0, source_row[0] - int(amount)), from_entity))
            
    # 4. Update demand status to 'Approved'
    cursor.execute('''
        UPDATE pol_demands SET status = 'Approved' WHERE id = ?
    ''', (demand_id,))
    
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/pol/allocate', methods=['POST'])
def allocate_pol():
    data = request.json
    from_entity = data.get("fromEntity")
    to_entity = data.get("toEntity")
    month = data.get("month")
    fiscal_year = data.get("fiscalYear", "2025-26")
    pol_grade = data.get("polGrade")
    amount = float(data.get("amount", 0))
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. Update pol_monthly_records for target unit (to_entity)
    cursor.execute('''
        SELECT id, allocation FROM pol_monthly_records 
        WHERE unit_name = ? AND fiscal_year = ? AND month = ? AND pol_grade = ?
    ''', (to_entity, fiscal_year, month, pol_grade))
    row = cursor.fetchone()
    if row:
        current_alloc = row[1] or 0.0
        cursor.execute('''
            UPDATE pol_monthly_records SET allocation = ? 
            WHERE id = ?
        ''', (current_alloc + amount, row[0]))
    else:
        cursor.execute('''
            INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
            VALUES (?, ?, ?, ?, ?, 0.0)
        ''', (to_entity, fiscal_year, month, pol_grade, amount))
        
    # 2. Update logistics.pol (on-hand inventory) for target unit
    cursor.execute("SELECT pol FROM logistics WHERE unit_name = ?", (to_entity,))
    target_row = cursor.fetchone()
    if target_row:
        cursor.execute('''
            UPDATE logistics SET pol = ? WHERE unit_name = ?
        ''', (target_row[0] + int(amount), to_entity))
    else:
        cursor.execute('''
            INSERT INTO logistics (unit_name, vAvail, vTotal, pol, cook, waiter, strength)
            VALUES (?, 0, 0, ?, 0, 0, 0)
        ''', (to_entity, int(amount)))
        
    # 3. Deduct from source unit (from_entity) if it's not 'Area HQ'
    if from_entity != 'Area HQ':
        cursor.execute("SELECT pol FROM logistics WHERE unit_name = ?", (from_entity,))
        source_row = cursor.fetchone()
        if source_row:
            cursor.execute('''
                UPDATE logistics SET pol = ? WHERE unit_name = ?
            ''', (max(0, source_row[0] - int(amount)), from_entity))
            
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/pol/expenditure', methods=['POST'])
def add_pol_expenditure():
    data = request.json
    unit_name = data.get("unitName")
    month = data.get("month")
    fiscal_year = data.get("fiscalYear", "2025-26")
    pol_grade = data.get("polGrade")
    amount = float(data.get("amount", 0))
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. Update pol_monthly_records expenditure for unit
    cursor.execute('''
        SELECT id, expenditure FROM pol_monthly_records 
        WHERE unit_name = ? AND fiscal_year = ? AND month = ? AND pol_grade = ?
    ''', (unit_name, fiscal_year, month, pol_grade))
    row = cursor.fetchone()
    if row:
        current_exp = row[1] or 0.0
        cursor.execute('''
            UPDATE pol_monthly_records SET expenditure = ? 
            WHERE id = ?
        ''', (current_exp + amount, row[0]))
    else:
        cursor.execute('''
            INSERT INTO pol_monthly_records (unit_name, fiscal_year, month, pol_grade, allocation, expenditure)
            VALUES (?, ?, ?, ?, 0.0, ?)
        ''', (unit_name, fiscal_year, month, pol_grade, amount))
        
    # 2. Update logistics.pol (on-hand inventory) for unit (deduct expenditure)
    cursor.execute("SELECT pol FROM logistics WHERE unit_name = ?", (unit_name,))
    target_row = cursor.fetchone()
    if target_row:
        cursor.execute('''
            UPDATE logistics SET pol = ? WHERE unit_name = ?
        ''', (max(0, target_row[0] - int(amount)), unit_name))
        
    conn.commit()
    conn.close()
    return jsonify({"success": True})


if __name__ == '__main__':
    init_db()
    # Server will run on port 3000 locally
    app.run(host='0.0.0.0', port=3000, debug=True)
